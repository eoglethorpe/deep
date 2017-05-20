import string
import json
from datetime import datetime

from excel_writer import ExcelWriter, RowCollection
from report.models import DisasterType, ReportStatus,\
                          HumanProfileField, PeopleInNeedField,\
                          HumanAccessField, HumanAccessPinField
from leads.models import Event
from openpyxl.styles import Font, Border, Side  # , Color


def get_dict(data, fields, default=''):
    _data = data
    fields = fields.split('.')
    for field in fields:
        if not isinstance(_data, dict):
            return _data
        _data = _data.get(field)
        if _data is None:
            return default
    return _data


def get_new_source(new):
    source = ''
    for s in new:
        date = s.get('date', None)
        if date:
            try:
                date = datetime.strptime(date, '%Y-%m-%d').strftime('%d-%m-%y')
            except:
                pass
        else:
            date = ''
        source += str(s.get('name')) + ' ' + date + '\n'
    return source.strip()


def get_source(data, fields, defailt=''):
    new = get_dict(data, fields + '.new', None)
    if new:
        return get_new_source(new)
    else:
        return get_dict(data, fields + '.old')


def map_day(num):
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday',
            'Friday', 'Saturday', 'Sunday']
    try:
        num = int(num) if isinstance(num, str) else num
        return days[num-1]
    except (ValueError, TypeError, IndexError):
        pass
    return ''


def export_xls(title):

    # Create a spreadsheet and get active workbook
    ew = ExcelWriter()
    ws = ew.get_active()
    ws.title = "Weekly Report"

    # Create title row
    titles = [
        "Crisis", "Country", "Week Covered By Report", "Day Covered By Report",
        "Type Of Disaster", "Status",
    ]

    hpf = HumanProfileField.objects.filter(parent__isnull=True)
    pinf = PeopleInNeedField.objects.filter(parent__isnull=True)
    haf = HumanAccessField.objects.all()
    hapf = HumanAccessPinField.objects.all()

    # HUMANITARIAN PROFILE
    for field in hpf:
        titles.append(field.name+'__Number__Source__Comment')
        for ch_field in field.humanprofilefield_set.all():
            titles.append(ch_field.name+'__Number__Source__Comment')

    # TODO: Replace Source/Date with Source
    # PEOPLE IN NEED
    for field in pinf:
        titles.append(
                field.name+'__'
                'Total__Source/Date__Comment__'
                'At Risk__Source/Date__Comment__'
                'Moderate__Source/Date__Comment__'
                'Severe__Source/Date__Comment__'
                'Planned__Source/Date__Comment')
        for ch_field in field.peopleinneedfield_set.all():
            titles.append(
                ch_field.name+'__'
                'Total__Source/Date__Comment__'
                'At Risk__Source/Date__Comment__'
                'Moderate__Source/Date__Comment__'
                'Severe__Source/Date__Comment__'
                'Planned__Source/Date__Comment')

    # ipc
    titles.append('ipc__None/Minimal__Stressed__Crisis__Emergency__Famine'
                  '__Source__Comment')

    # HUMANITARIAN ACCESS
    for field in haf:
        titles.append(field.name+'__Yes/No__Source/Date__Comment')

    # HUMANITARIAN ACCESS PIN
    for field in hapf:
        titles.append(field.name+'__Number__Source/Date__Comment')

    # Create Columns
    col = 0
    for t in titles:
        i = col
        splits = t.split('__')

        ws.cell(row=1, column=i+1).value = splits[0]
        ws.cell(row=1, column=i+1).font = Font(bold=True)

        if len(splits) > 1:
            for j, split in enumerate(splits[1:]):
                ws.cell(row=2, column=i+j+1).value = split
                ws.cell(row=2, column=i+j+1).font = Font(bold=True)

            col = col + len(splits) - 2
            ws.merge_cells(start_row=1, end_row=1,
                           start_column=i+1, end_column=i+len(splits)-1)

        col = col + 1

    ew.auto_fit_cells_in_row(1, ws)
    ew.auto_fit_cells_in_row(2, ws)

    # Thick, dotted border for readibility
    for cell in list(ws.rows)[1-1]:
            cell.border = Border(
                bottom=Side(border_style='dotted'))
    for cell in list(ws.rows)[2-1]:
            cell.border = Border(
                bottom=Side(border_style='thick'))

    # Create Rows

    for event in Event.objects.all():
        for report in event.weeklyreport_set.all():
            rows = RowCollection(1)
            data = json.loads(report.data)

            # Report Info
            isoreport = report.start_date.isocalendar()

            row = [
                event.name, report.country.name,
                'Week ' + str(isoreport[1]) + ' ' + str(isoreport[0]),
                map_day(data.get('day-select')),

                DisasterType.objects.get(
                    pk=get_dict(data, 'disaster_type')).name
                if get_dict(data, 'disaster_type') else '',

                ReportStatus.objects.get(pk=get_dict(data, 'status'))
                if get_dict(data, 'status') else '',
            ]

            # HUMANITARIAN PROFILE
            for field in hpf:
                row.extend([
                    get_dict(data, 'human.number.'+str(field.pk)),
                    get_source(data, 'human.source.'+str(field.pk)),
                    get_dict(data, 'human.comment.'+str(field.pk))
                ])

                for ch_field in field.humanprofilefield_set.all():
                    row.extend([
                        get_dict(data, 'human.number.'+str(ch_field.pk)),
                        get_source(data, 'human.source.' +
                                 str(ch_field.pk)),
                        get_dict(data, 'human.comment.'+str(ch_field.pk))
                    ])

            # PEOPLE IN NEED
            for field in pinf:
                _fields = ['total', 'at-risk', 'moderate',
                           'severe', 'planned']
                _data = []
                for _field in _fields:
                    _data.extend([
                        get_dict(data, 'people.'+_field+'.'+str(field.pk)),
                        get_source(data, 'people.'+_field +
                                 '-source.'+str(field.pk)),
                        get_dict(data, 'people.'+_field +
                                 '-comment.'+str(field.pk)),
                        ])
                row.extend(_data)
                for ch_field in field.peopleinneedfield_set.all():
                    _data = []
                    for _field in _fields:
                        _data.extend([
                            get_dict(data, 'people.'+_field +'.'+
                                     str(ch_field.pk)),
                            get_source(data, 'people.'+_field +
                                     '-source.'+str(ch_field.pk)),
                            get_dict(data, 'people'+_field +
                                     '-comment'+str(ch_field.pk)),
                            ])
                    row.extend(_data)

            # ipc
            row.extend([data['ipc'][ipc_field] for ipc_field in 'abcde'])
            row.extend([get_source(data, 'ipc.f')])
            row.extend([data['ipc']['g']])

            # HUMANITARIAN ACCESS
            for field in haf:
                row.extend([
                    get_dict(data, 'access.'+str(field.pk)),
                    get_source(data, 'access-extra.source.' +
                             str(field.pk)),
                    get_dict(data, 'access-extra.comment.'+str(field.pk)),
                    ])

            # HUMANITARIAN ACCESS PIN
            for field in hapf:
                row.extend([
                    get_dict(data, 'access-pin.number.'+str(field.pk)),
                    get_source(data, 'access-pin.source.'+str(field.pk)),
                    get_dict(data, 'access-pin.comment.'+str(field.pk)),
                    ])
            #                   Add to Row Collection
            rows.add_values(row)

            #                   Add to Workbench
            ew.append(rows.rows, ws)

    # ew.wb.save('balances.xlsx')
    return ew.get_http_response(title)


def xstr(conv):
    """remove illegal characters from a string (errors from PDFs etc)"""
    return "".join(filter(lambda x: x in string.printable, conv))
