from django.core.management.base import BaseCommand, CommandError
from django.contrib.staticfiles import finders
from django.core.files import File
from django.db.models import Q

from openpyxl import load_workbook
from openpyxl import Workbook

import os
import glob

from deep.settings import BASE_DIR
from leads.models import *
from entries.models import *
from entries.export_entries_xls import *

class Command(BaseCommand):
    def handle(self, *args, **kwargs):
        folder = os.path.join(BASE_DIR, 'static/global_export/')

        new_wb = Workbook()
        first = True

        files = glob.glob(os.path.join(folder, '*.xlsx'))
        for i, file in enumerate(files):
            if 'all.xlsx' in file:
                continue
            print('{}%     '.format(int(i/len(files)*100)), end='\r')

            wb = load_workbook(file)
            for sheet in wb:
                if sheet.title not in new_wb.sheetnames:
                    # New sheet
                    if first:
                        new_sheet = new_wb.active
                        new_sheet.title = sheet.title
                        first = False
                    else:
                        new_sheet = new_wb.create_sheet(sheet.title)

                    # Copy header cells
                    for row in sheet.iter_rows(min_row=1, max_row=1):
                        new_sheet.append([ cell.value for cell in row ])

                new_sheet = new_wb.get_sheet_by_name(sheet.title)
                for row in sheet.iter_rows(min_row=2):
                    new_sheet.append([ cell.value for cell in row ])

        new_wb.save(os.path.join(folder, 'all.xlsx'))
