#TODO: rewrite export_files to except **kwargs and then merge gen bvs and ias

"""exporting entries to xlsx"""
import itertools
from copy import deepcopy

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font

from entries.export_fields import *
from entries.models import *

GROUPED_TAB = 'Grouped Entries'
SPLIT_TAB = 'Split Entries'
META_TAB =  'Metadata'

def _expand_cols(wb):
    """expand all column dims to match max length of header"""
    for sht in wb.worksheets:
        for col in sht.columns:
            if col[0].value:
                sht.column_dimensions[col[0].column].width = max(len(str(col[0].value)), 30)

            #the below makes cells taller
            #cell.alignment = Alignment(wrap_text=True, vertical='top')

def _fill_cells(wb):
    """cheeky fix to prevent overflowing cells - add a space"""
    for sht in wb.worksheets:
        for c in list(itertools.chain(*sht.rows)):
            if c.value in (None, ''):
                c.value = ' '

def _add_filters(ws):
    """create filters for every column"""
    for i,c in enumerate(ws.columns):
         ws.auto_filter.ref = ws

def init_xls():
    #initialize Workbook and activie sheet
    wb = Workbook()
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    wb.create_sheet(SPLIT_TAB)
    wb.create_sheet(GROUPED_TAB)
    wb.create_sheet(META_TAB)

    return wb

def _make_bold(cells):
    """make a list of cells bold"""
    for c in cells:
        c.font = Font(bold = True)

def _mk_col_nm(cn):
    return cn.replace(' ', '_').lower()

def _mk_col_nm_lst(cl):
    return [_mk_col_nm(cn) for cn in cl]

def _gen_base_vals():
    """which columns should be included and which function is used to create them"""
    return OrderedDict([
            ('Date of Lead Publication' , 'get_lead_created_at_tm'),
            ('Date of Entry Creation' , 'get_ent_created_at_m'),
            ('Created By' , 'get_created_by'),
            ('Lead Title' , 'get_lead_nm'),
            ('Confidentiality', 'get_confidentiality'),
            ('Source', 'get_source'),
            ('Crisis Name', 'get_crisis'),
            ('Country(s)', 'get_countries'),
            ('Admin Level 1', 'get_admn_lvl1'),
            ('Admin Level 2', 'get_admn_lvl2'),
            ('Admin Level 3', 'get_admn_lvl3'),
            ('Admin Level 4', 'get_admn_lvl4'),
            ('Admin Level 5', 'get_admn_lvl5'),
            ('Affected Groups Level 1', 'get_aff_lvl1'),
            ('Affected Groups Level 2', 'get_aff_lvl2'),
            ('Affected Groups Level 3', 'get_aff_lvl3'),
            ('Vulnerable Groups' , 'get_vuln_list'),
            ('Specific Needs Groups' , 'get_specific_list')])

def _gen_ias():
    """similar to base_vals but for ias"""
    return OrderedDict([
            ('Information Attribute Level 1', 'get_ia_lvl1'),
            ('Information Attribute Level 2', 'get_ia_lvl2'),
            ('Excerpt', 'get_ia_exc'),
            ('Number', 'get_ia_num'),
            ('Reliability', 'get_ia_rel'),
            ('Severity', 'get_ia_sev')])

def _gen_ids():
    """work around to put IDs at the end"""
    return OrderedDict([
            ('Lead ID', 'get_lead_id'),
            ('Entry ID', 'get_entry_id'),
            ('Tag ID', 'get_tag_id')])

def _split_row(row):
    """a crude method for breaking up a row"""

    #used to flag if a list returned by recusrive fun shouldn't be included (work around)
    class skip():
        pass

    out = []
    #a recursive function that creates permutations of rows based on their contents that are lists
    def rec(l, it):
        if it == len(l):
            return []

        elif isinstance(l[it], list):
            if len(l[it]) > 0:
                for v in l[it]:
                    tmp = deepcopy(l)
                    tmp[it] = v
                    out.append((rec(tmp, 0)))

                return [skip]

            else:
                tmp = deepcopy(l)
                tmp[it] = ''
                out.append((rec(tmp, 0)))
                return [skip]


        else:
            return [l[it]] + rec(l,it+1)

    #only do recurisve if our row contains > 1 list, else return just the row. realistically it always will
    if True in [isinstance(v, list) for v in row]:
        rec(row, 0)
        l = [r for r in out if skip not in r]
        return [r for r in out if skip not in r]
    else:
        return [row]

def _gen_out(sht, type):
    """Generate either a split or grouped sheet"""

    #values of dicts are method names to be used in lookups
    ents = Entry.objects.all()
    base_cols = _gen_base_vals()
    ia_cols = _gen_ias()
    id_cols =  _gen_ids()

    #insert header for base cols and IA cols
    sht.append(list(base_cols.keys()) + list(ia_cols.keys()) + list(id_cols.keys()))

    _make_bold(sht.rows[0])

    for e in ents:
        for att in e.attributedata_set.all():
            out = [globals()[t](e) for t in base_cols.values()] + \
                            [globals()[t](att) for t in ia_cols.values()] + \
                            [globals()[t](e, att) for t in id_cols.values()]

            if type == GROUPED_TAB:
                sht.append([', '.join(v) if isinstance(v, list) else v for v in out])
            elif type == SPLIT_TAB:
                for v in _split_row(out):
                    sht.append(v)
            else:
                raise Exception('Must specify output type')

def _gen_meta(sht):
    """return a sheet with metadata info on export"""
    sht.append(['Export Information'])
    sht.append(['Date', time.strftime("%Y-%m-%d at %H:%M")])
    sht.append(['Number of entries', len(Entry.objects.all())])

    _make_bold(sht.rows[0])


def export():
    wb = init_xls()
    _gen_out(wb.get_sheet_by_name(SPLIT_TAB), SPLIT_TAB)
    _gen_out(wb.get_sheet_by_name(GROUPED_TAB), GROUPED_TAB)
    _gen_meta(wb.get_sheet_by_name(META_TAB))

    _expand_cols(wb)
    _fill_cells(wb)
    _add_filters(wb.get_sheet_by_name(GROUPED_TAB))
    return save_virtual_workbook(wb)