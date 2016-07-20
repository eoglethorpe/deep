"""exporting entries to xlsx or docx with a given heirarchy"""
import time
import os

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Font

from entries.models import *

EXPORT_TAB = 'DEEP Export | Entries'
META_TAB =  'Metadata'

def _expand_cols(wb):
    """expand all column dims to match max length"""
    for sht in wb.worksheets:
        dims = {}
        for row in sht.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column] = max((dims.get(cell.column, 0), len(cell.value)))
        for col, value in dims.items():
            sht.column_dimensions[col].width = value + 1

def init_xls():
    #initialize Workbook and activie sheet
    wb = Workbook()
    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    wb.create_sheet(EXPORT_TAB, 0)
    wb.create_sheet(META_TAB)

    return wb

def gen_meta(sht):
    """return a sheet with metadata info on export"""
    sht.append(['Export Information'])
    sht.append(['Date', time.strftime("%Y-%m-%d at %H:%M")])
    sht.append(['Number of entries', 'todo'])

    _make_bold(sht.rows[0])

def _make_bold(cells):
    """make a list of cells bold"""
    for c in cells:
        c.font = Font(bold = True)

def gen_exports(sht):
    COLS = ['Affected Groups', 'Created At', 'Created By', 'Lead', 'Map Selections']
    sht.append(COLS)
    _make_bold(sht.rows[0])

    for e in Entry.objects.all():
        sht.append([
            str(e.affected_groups.all()),
            "{0:%b %d %Y %I:%M%p}".format(e.created_at),
            e.created_by.username,
            e.lead.name,
            str(e.map_selections)])

def export_xls():
    wb = init_xls()
    gen_exports(wb.get_sheet_by_name(EXPORT_TAB))
    gen_meta(wb.get_sheet_by_name(META_TAB))

    _expand_cols(wb)
    return save_virtual_workbook(wb)


def main(type):
    if type == 'xls':
        return export_xls()